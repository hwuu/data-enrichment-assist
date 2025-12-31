"""GaussDB Operations Ticket Viewer - FastAPI App"""
import io
from datetime import datetime
from pathlib import Path
from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from config import DATABASE_CONFIG, SERVER_HOST, SERVER_PORT, TICKET_URL_PATTERN
from database import create_database

app = FastAPI(title="GaussDB Ops Viewer", description="运维工单浏览器")

BASE_DIR = Path(__file__).resolve().parent
templates = Jinja2Templates(directory=BASE_DIR / "templates")

# Configure Jinja2 to not escape unicode in tojson
templates.env.policies['json.dumps_kwargs'] = {'ensure_ascii': False}

# Initialize database
db = create_database(DATABASE_CONFIG)


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    """Main page - loads only ticket summaries."""
    tickets = db.get_ticket_list()

    # Extract unique values for filters
    issue_types = sorted(set(t['issueType'] for t in tickets))
    owners = sorted(set(t['owner'] for t in tickets))

    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "tickets": tickets,
            "issue_types": issue_types,
            "owners": owners,
            "ticket_url_pattern": TICKET_URL_PATTERN
        }
    )


@app.get("/api/tickets")
async def api_tickets():
    """API endpoint for tickets."""
    return db.get_all_tickets()


@app.get("/api/tickets/{process_id}")
async def api_ticket_detail(process_id: str):
    """API endpoint for single ticket."""
    ticket = db.get_ticket_by_id(process_id)
    if ticket:
        return ticket
    return {"error": "not found"}


@app.get("/api/tickets/{process_id}/review")
async def api_get_review(process_id: str):
    """Get review for a ticket."""
    review = db.get_ticket_review(process_id)
    if review:
        return review
    return {"processId": process_id, "conclusion": None, "content": "", "createTime": None, "updateTime": None}


@app.post("/api/tickets/{process_id}/review")
async def api_save_review(process_id: str, request: Request):
    """Save review for a ticket."""
    body = await request.json()
    conclusion = body.get("conclusion", "")
    content = body.get("content", "")
    review = db.save_ticket_review(process_id, conclusion, content)
    return review


@app.get("/api/export")
async def api_export(
    type: str = "all",
    owner: str = "all",
    score: str = "all",
    review: str = "all",
    sort: str = "updateTime-desc"
):
    """Export tickets to Excel with filters and sorting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo

    tickets = db.get_ticket_list()

    # Apply filters
    filtered = []
    for ticket in tickets:
        if type != "all" and ticket["issueType"] != type:
            continue
        if owner != "all" and ticket["owner"] != owner:
            continue
        if score != "all":
            s = ticket["score"]
            if score == "high" and s < 8:
                continue
            if score == "medium" and (s < 6 or s >= 8):
                continue
            if score == "low" and s >= 6:
                continue
        if review != "all":
            if review == "pending":
                if ticket.get("hasReview"):
                    continue
            elif review == "expired":
                if not ticket.get("reviewExpired"):
                    continue
            else:
                # 通过/不通过/待定
                if ticket.get("reviewExpired") or ticket.get("conclusion") != review:
                    continue
        filtered.append(ticket)

    # Apply sorting
    sort_field, sort_order = sort.split("-") if "-" in sort else (sort, "desc")
    reverse = sort_order == "desc"

    if sort_field == "id":
        filtered.sort(key=lambda t: t["processId"], reverse=reverse)
    elif sort_field == "score":
        filtered.sort(key=lambda t: t["score"] or 0, reverse=reverse)
    elif sort_field == "createTime":
        filtered.sort(key=lambda t: t["createTime"] or "", reverse=reverse)
    elif sort_field == "updateTime":
        filtered.sort(key=lambda t: t["updateTime"] or "", reverse=reverse)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "工单列表"

    # Headers
    headers = ["工单ID", "URL", "问题类型", "负责人", "问题描述", "得分", "审核结论", "审核意见"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Get all reviews in one query (avoid N+1 problem)
    all_reviews = db.get_all_reviews()

    # Data rows
    for row_idx, ticket in enumerate(filtered, 2):
        # Generate URL if pattern is set
        url = ""
        if TICKET_URL_PATTERN:
            url = TICKET_URL_PATTERN.replace("{processId}", ticket["processId"])

        # Get review info from cached dict
        rev = all_reviews.get(ticket["processId"])
        conclusion = rev.get("conclusion", "") if rev else ""
        content = rev.get("content", "") if rev else ""

        ws.cell(row=row_idx, column=1, value=ticket["processId"])
        url_cell = ws.cell(row=row_idx, column=2)
        if url:
            url_cell.value = "🔗"
            url_cell.hyperlink = url
            url_cell.font = Font(color="0563C1")
            url_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=ticket["issueType"])
        ws.cell(row=row_idx, column=4, value=ticket["owner"])
        ws.cell(row=row_idx, column=5, value=ticket["problem"])
        ws.cell(row=row_idx, column=6, value=ticket["score"])
        ws.cell(row=row_idx, column=7, value=conclusion)
        ws.cell(row=row_idx, column=8, value=content)

    # Adjust column widths
    # Fixed width for 问题描述(E) and 审核意见(H): 80 characters
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["H"].width = 80

    # Auto-fit other columns based on content
    for col_letter in ["A", "B", "C", "D", "F", "G"]:
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
        ws.column_dimensions[col_letter].width = max(max_length + 2, 4)

    # Create table for filter/sort in Excel
    if filtered:
        table_range = f"A1:H{len(filtered) + 1}"
        table = Table(displayName="TicketTable", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(table)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"tickets_export_{timestamp}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host=SERVER_HOST, port=SERVER_PORT)
